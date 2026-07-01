from flask import render_template, redirect, url_for, session, request, flash
from . import main_bp
import mysql.connector
from mysql.connector import Error
import hashlib
from functools import wraps

# Importar funciones generales
from app.funciones.funGeneral import (
    get_db_connection,
    hash_password,
    get_departamentos as get_departamentos_func,
    get_provincias as get_provincias_func,
    get_distritos as get_distritos_func,
    get_roles as get_roles_func,
    get_areas as get_areas_func,
    get_cargos as get_cargos_func,
    get_fuentes_contacto_api as get_fuentes_contacto_api_func,
    get_proyectos_api as get_proyectos_api_func,
    get_estados_prospeccion_api as get_estados_prospeccion_api_func,
    get_tipos_compra_api as get_tipos_compra_api_func
)

# Importar funciones de usuarios/asesores
from app.funciones.register_user import (
    register_asesor_api as register_asesor_api_func,
    get_usuarios_asesores as get_usuarios_asesores_func,
    get_asesor_by_documento as get_asesor_by_documento_func,
    update_asesor_api as update_asesor_api_func,
    delete_asesor_api as delete_asesor_api_func,
    cambiar_contrasena_api as cambiar_contrasena_api_func,
    buscar_usuario_por_documento as buscar_usuario_por_documento_func
)

# Importar funciones de clientes
from app.funciones.clientes import (
    insertar_cliente_api,
    listar_clientes_api,
    listar_todos_clientes_api,
    eliminar_cliente_api,
    obtener_cliente_por_documento_api,
    actualizar_cliente_api,
    descartar_cliente_api,
    contar_clientes_por_estado_prospeccion_api,
    contar_clientes_descartados_api,
    registrar_seguimiento_api,
    listar_seguimientos_cliente_api,
    listar_tipos_seguimiento_api,
    listar_historial_seguimientos_api,
    listar_ultimos_3_seguimientos_api,
    filtrar_clientes_api,
    importar_clientes_api
)

# Decorador para requerir autenticación
def login_required(f):
    """Decorador para proteger rutas que requieren autenticación"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_email' not in session:
            flash('Debes iniciar sesión para acceder a esta página', 'warning')
            return redirect(url_for('auth.login'))
        return f(*args, **kwargs)
    return decorated_function

@main_bp.route('/')
def index():
    """Ruta raíz - redirige a bienvenida si no está autenticado"""
    if 'user_email' in session:
        return redirect(url_for('main.dashboard'))
    return redirect(url_for('auth.welcome'))

@main_bp.route('/dashboard')
@login_required
def dashboard():
    return render_template('dashboard.html')

@main_bp.route('/clients')
@login_required
def clients():
    """Página de gestión de clientes"""
    return render_template('clients.html')

@main_bp.route('/simulador')
@login_required
def simulador():
    """Página de simulador de ventas de terrenos"""
    return render_template('simulador.html')



@main_bp.route('/register-user', methods=['GET', 'POST'])
@login_required
def register_user():
    """Página para registrar nuevos usuarios del sistema"""
    if request.method == 'POST':
        # Obtener datos del formulario
        num_documento = request.form.get('num_documento')
        nombres = request.form.get('nombres')
        apellido_paterno = request.form.get('apellido_paterno')
        apellido_materno = request.form.get('apellido_materno')
        email = request.form.get('email')
        celular = request.form.get('celular')
        numero_emergencia = request.form.get('numero_emergencia')
        usuario = request.form.get('usuario')
        password = request.form.get('password')
        id_rol = request.form.get('id_rol')  # Cambiar a id_rol
        
        # Validar que todos los campos estén completos
        if not all([num_documento, nombres, apellido_paterno, email, usuario, password, id_rol]):
            flash('Por favor completa todos los campos requeridos', 'error')
            return render_template('register_user.html')
        
        connection = get_db_connection()
        if not connection:
            flash('Error de conexión con la base de datos', 'error')
            return render_template('register_user.html')
        
        try:
            cursor = connection.cursor(dictionary=True)
            
            # Verificar si el usuario ya existe
            query_check = "SELECT usuario FROM TblUsuarios WHERE usuario = %s"
            cursor.execute(query_check, (usuario,))
            if cursor.fetchone():
                flash('El usuario ya existe en el sistema', 'error')
                cursor.close()
                connection.close()
                return render_template('register_user.html')
            
            # Verificar si el documento ya existe en TblPersona
            query_check_doc = "SELECT num_documento FROM TblPersona WHERE num_documento = %s"
            cursor.execute(query_check_doc, (num_documento,))
            if cursor.fetchone():
                flash('Este número de documento ya está registrado', 'error')
                cursor.close()
                connection.close()
                return render_template('register_user.html')
            
            # Insertar en TblPersona
            query_persona = """
                INSERT INTO TblPersona (
                    num_documento, nombres, apellido_paterno, apellido_materno,
                    email, celular, numero_emergencia, fecha_creacion, estado
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, NOW(), 'Activo')
            """
            cursor.execute(query_persona, (
                num_documento, nombres, apellido_paterno, apellido_materno,
                email, celular, numero_emergencia
            ))
            
            # Encriptar contraseña
            password_hash = hash_password(password)
            
            # Insertar en TblUsuarios con id_rol (sin campo 'rol' redundante)
            query_usuario = """
                INSERT INTO TblUsuarios (
                    num_documento, usuario, password_hash, id_rol, estado,
                    fecha_creacion, ultimo_acceso, intentos_fallidos
                ) VALUES (%s, %s, %s, %s, 'Activo', NOW(), NULL, 0)
            """
            cursor.execute(query_usuario, (
                num_documento, usuario, password_hash, id_rol
            ))
            
            connection.commit()
            flash(f'¡Usuario {usuario} registrado exitosamente!', 'success')
            cursor.close()
            connection.close()
            return redirect(url_for('main.register_user'))
            
        except Error as e:
            print(f"Error en registro: {e}")
            flash('Error al registrar el usuario', 'error')
            connection.close()
            return render_template('register_user.html')
    
    return render_template('register_user.html')


@main_bp.route('/api/departamentos', methods=['GET'])
@login_required
def get_departamentos():
    """Obtener lista de departamentos"""
    return get_departamentos_func()


@main_bp.route('/api/provincias/<int:id_departamento>', methods=['GET'])
@login_required
def get_provincias(id_departamento):
    """Obtener lista de provincias por departamento"""
    return get_provincias_func(id_departamento)


@main_bp.route('/api/distritos/<int:id_provincia>', methods=['GET'])
@login_required
def get_distritos(id_provincia):
    """Obtener lista de distritos por provincia"""
    return get_distritos_func(id_provincia)


@main_bp.route('/api/register-asesor', methods=['POST'])
@login_required
def register_asesor_api():
    """API para registrar un nuevo asesor desde el modal"""
    return register_asesor_api_func()


@main_bp.route('/api/usuarios-asesores', methods=['GET'])
@login_required
def get_usuarios_asesores():
    """Obtener listado de usuarios asesores con datos personales"""
    return get_usuarios_asesores_func()


@main_bp.route('/api/roles', methods=['GET'])
@login_required
def get_roles():
    """Obtener listado de roles disponibles"""
    return get_roles_func()


@main_bp.route('/api/areas', methods=['GET'])
@login_required
def get_areas():
    """Obtener listado de áreas disponibles"""
    return get_areas_func()


@main_bp.route('/api/cargos', methods=['GET'])
@main_bp.route('/api/cargos/<int:id_area>', methods=['GET'])
@login_required
def get_cargos(id_area=None):
    """Obtener listado de cargos disponibles, opcionalmente filtrados por área"""
    return get_cargos_func(id_area)


@main_bp.route('/api/buscar-usuario', methods=['GET'])
@login_required
def buscar_usuario_por_documento():
    """Buscar usuario por documento (para validación de jerarquía)"""
    return buscar_usuario_por_documento_func()


@main_bp.route('/api/asesor/<num_documento>', methods=['GET'])
@login_required
def get_asesor_by_documento(num_documento):
    """Obtener datos de un asesor específico para edición"""
    return get_asesor_by_documento_func(num_documento)


@main_bp.route('/api/update-asesor', methods=['POST'])
@login_required
def update_asesor_api():
    """API para actualizar un asesor existente"""
    return update_asesor_api_func()

@main_bp.route('/api/delete-asesor/<num_documento>', methods=['DELETE'])
@login_required
def delete_asesor_api(num_documento):
    """API para eliminar un asesor de todas las tablas"""
    return delete_asesor_api_func(num_documento)


@main_bp.route('/api/cambiar-contrasena', methods=['POST'])
@login_required
def cambiar_contrasena_api():
    """API para cambiar contraseña del usuario actual"""
    return cambiar_contrasena_api_func()


# ============================================================================
# RUTAS API PARA GESTIÓN DE CLIENTES
# ============================================================================

@main_bp.route('/api/fuentes-contacto', methods=['GET'])
@login_required
def get_fuentes_contacto():
    """API para obtener lista de fuentes de contacto"""
    return get_fuentes_contacto_api_func()


@main_bp.route('/api/proyectos', methods=['GET'])
@login_required
def get_proyectos():
    """API para obtener lista de proyectos"""
    return get_proyectos_api_func()


@main_bp.route('/api/estados-prospeccion', methods=['GET'])
@login_required
def get_estados_prospeccion():
    """API para obtener lista de estados de prospeccion"""
    return get_estados_prospeccion_api_func()


@main_bp.route('/api/tipos-compra', methods=['GET'])
@login_required
def get_tipos_compra():
    """API para obtener lista de tipos de compra"""
    return get_tipos_compra_api_func()

# ============================================================================
# RUTAS PARA IMPORTACIÓN Y DESCARGA DE PLANTILLA DE CLIENTES
# ============================================================================

@main_bp.route('/api/clientes/descargar-plantilla', methods=['GET'])
def descargar_plantilla_clientes():
    """Descargar plantilla Excel para importar clientes"""
    import os
    from flask import send_file
    from datetime import datetime
    
    # Generar nombre de archivo
    filename = f'Plantilla_Clientes_{datetime.now().strftime("%d_%m_%Y")}.xlsx'
    filepath = os.path.join(os.path.dirname(__file__), '..', 'static', 'templates', filename)
    
    # Si el archivo ya existe y es del mismo día, usarlo
    if os.path.exists(filepath):
        return send_file(
            filepath,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    # Si no existe, generar uno nuevo
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import io
    
    # Crear directorio si no existe
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    # Crear nuevo workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Clientes"
    
    # Definir estilos
    header_fill = PatternFill(start_color="4D148C", end_color="4D148C", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # Definir encabezados
    headers = [
        "Tipo Documento*",
        "Nro. Documento*",
        "Género*",
        "Nombres*",
        "Apellido Paterno*",
        "Apellido Materno",
        "Fecha Nacimiento*",
        "Estado Civil*",
        "Email*",
        "Celular*",
        "Dirección*",
        "Departamento*",
        "Provincia*",
        "Distrito*",
        "Fuente Contacto",
        "Proyecto Interés",
        "Estado Prospección",
        "Tipo Compra",
        "Prioridad"
    ]
    
    # Escribir encabezados
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border
    
    # Ajustar ancho de columnas
    column_widths = [18, 18, 12, 18, 20, 20, 18, 15, 25, 18, 30, 18, 18, 18, 18, 20, 20, 18, 12]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # Agregar fila de ejemplo
    example_data = [
        "DNI",
        "12345678",
        "M",
        "Juan",
        "Pérez",
        "García",
        "1990-01-15",
        "Casado",
        "juan@ejemplo.com",
        "+51 999 999 999",
        "Av. Principal 123, Lima",
        "Lima",
        "Lima",
        "Lima",
        "",
        "Villa de los Nísperos",
        "POTENCIAL",
        "Contado",
        "Media"
    ]
    
    for col, value in enumerate(example_data, 1):
        cell = ws.cell(row=2, column=col)
        cell.value = value
        cell.border = border
        if col in [2]:
            cell.alignment = center_alignment
    
    # ============================================================================
    # CREAR SEGUNDA HOJA CON INSTRUCCIONES Y OPCIONES
    # ============================================================================
    ws_instrucciones = wb.create_sheet("Instrucciones")
    
    # Título
    title_cell = ws_instrucciones['A1']
    title_cell.value = "GUÍA DE IMPORTACIÓN DE CLIENTES"
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="4D148C", end_color="4D148C", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_instrucciones.merge_cells('A1:D1')
    ws_instrucciones.row_dimensions[1].height = 25
    
    # Sección: Instrucciones Generales
    row = 3
    header_fill_light = PatternFill(start_color="E6D5F0", end_color="E6D5F0", fill_type="solid")
    header_font_dark = Font(bold=True, color="4D148C", size=11)
    
    # Instrucciones
    ws_instrucciones[f'A{row}'] = "INSTRUCCIONES GENERALES"
    ws_instrucciones[f'A{row}'].font = header_font_dark
    ws_instrucciones[f'A{row}'].fill = header_fill_light
    row += 1
    
    instrucciones = [
        "1. Los campos marcados con (*) son OBLIGATORIOS",
        "2. No agregues más columnas de las que aparecen en la plantilla",
        "3. Verifica que todos los valores coincidan con las opciones válidas",
        "4. Las fechas deben estar en formato: YYYY-MM-DD (ej: 2026-01-15)",
        "5. Los teléfonos pueden incluir código de país (ej: +51 999 999 999)",
        "6. No dejes filas en blanco en el medio del documento",
        "7. Verifica que NO haya espacios en blanco al inicio/final de los campos"
    ]
    
    for instr in instrucciones:
        ws_instrucciones[f'A{row}'] = instr
        ws_instrucciones.row_dimensions[row].height = 18
        row += 1
    
    # Sección: Opciones por Campo
    row += 2
    ws_instrucciones[f'A{row}'] = "OPCIONES VÁLIDAS POR CAMPO"
    ws_instrucciones[f'A{row}'].font = header_font_dark
    ws_instrucciones[f'A{row}'].fill = header_fill_light
    row += 1
    
    # Tabla de opciones
    ws_instrucciones.column_dimensions['A'].width = 25
    ws_instrucciones.column_dimensions['B'].width = 50
    
    opciones = [
        ("TIPO DOCUMENTO", "DNI, RUC, PASAPORTE, CARNET EXTRANJERIA"),
        ("GÉNERO", "M (Masculino), F (Femenino)"),
        ("ESTADO CIVIL", "Soltero, Casado, Divorciado, Viudo, Conviviente"),
        ("DEPARTAMENTO", "Lima, Arequipa, Cajamarca, La Libertad, Piura, Etc."),
        ("ESTADO PROSPECCIÓN", "POTENCIAL, CALIENTE, FRIO, TRATAMIENTO"),
        ("TIPO COMPRA", "Contado, Financiado, Crédito, Otro"),
        ("PRIORIDAD", "Baja, Media, Alta"),
        ("FUENTE CONTACTO", "Referencia, Derivado, Internet, Publicidad, Otro"),
        ("PROYECTO INTERÉS", "Villa de los Nísperos, Las Arenas, Otro Proyecto"),
    ]
    
    for campo, valores in opciones:
        cell_campo = ws_instrucciones[f'A{row}']
        cell_campo.value = campo
        cell_campo.font = Font(bold=True, color="4D148C")
        cell_campo.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        cell_campo.border = border
        
        cell_valores = ws_instrucciones[f'B{row}']
        cell_valores.value = valores
        cell_valores.border = border
        cell_valores.alignment = Alignment(wrap_text=True, vertical="center")
        
        ws_instrucciones.row_dimensions[row].height = 30
        row += 1
    
    # Sección: Ejemplos
    row += 2
    ws_instrucciones[f'A{row}'] = "EJEMPLO DE DATOS VÁLIDOS"
    ws_instrucciones[f'A{row}'].font = header_font_dark
    ws_instrucciones[f'A{row}'].fill = header_fill_light
    row += 1
    
    ejemplo_texto = """
Tipo Documento: DNI
Nro. Documento: 12345678 (8 dígitos para DNI)
Género: M
Nombres: Juan
Apellido Paterno: Pérez
Apellido Materno: García (opcional)
Fecha Nacimiento: 1990-01-15
Estado Civil: Casado
Email: juan.perez@email.com
Celular: +51 999 999 999
Dirección: Av. Principal 123, Apto 402
Departamento: Lima
Provincia: Lima
Distrito: Lima
Fuente Contacto: Referencia
Proyecto Interés: Villa de los Nísperos
Estado Prospección: POTENCIAL
Tipo Compra: Contado
Prioridad: Media
    """
    
    ws_instrucciones[f'A{row}'] = ejemplo_texto
    ws_instrucciones[f'A{row}'].alignment = Alignment(wrap_text=True, vertical="top")
    ws_instrucciones.row_dimensions[row].height = 120
    
    # Ajustar anchos
    ws_instrucciones.column_dimensions['A'].width = 30
    ws_instrucciones.column_dimensions['B'].width = 60
    
    # Guardar archivo en disk
    wb.save(filepath)
    
    # Enviar archivo
    return send_file(
        filepath,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@main_bp.route('/api/clientes/importar', methods=['POST'])
@login_required
def importar_clientes():
    """Importar clientes desde archivo Excel"""
    from app.funciones.clientes import importar_clientes_api
    from flask import jsonify
    
    resultado, status_code = importar_clientes_api()
    return jsonify(resultado), status_code


# ============================================================================
# RUTAS CRUD PARA GESTIÓN DE CLIENTES
# ============================================================================

@main_bp.route('/api/clientes', methods=['GET'])
@login_required
def get_clientes():
    """Obtener lista de clientes con filtros"""
    return filtrar_clientes_api()


@main_bp.route('/api/clientes', methods=['POST'])
@login_required
def crear_cliente_api():
    """Crear un nuevo cliente"""
    return insertar_cliente_api()


@main_bp.route('/api/clientes/<num_documento>', methods=['GET'])
@login_required
def obtener_cliente(num_documento):
    """Obtener datos de un cliente específico"""
    return obtener_cliente_por_documento_api(num_documento)


@main_bp.route('/api/clientes/<num_documento>', methods=['PUT'])
@login_required
def actualizar_cliente_route(num_documento):
    """Actualizar datos de un cliente"""
    return actualizar_cliente_api(num_documento)


@main_bp.route('/api/clientes/<num_documento>', methods=['DELETE'])
@login_required
def eliminar_cliente_route(num_documento):
    """Eliminar un cliente"""
    return eliminar_cliente_api(num_documento)


@main_bp.route('/api/clientes/<num_documento>/descartar', methods=['PUT'])
@login_required
def descartar_cliente_route(num_documento):
    """Cambiar estado de cliente a descartado"""
    return descartar_cliente_api(num_documento)


# ============================================================================
# RUTAS PARA SEGUIMIENTOS DE CLIENTES
# ============================================================================

@main_bp.route('/api/clientes/<num_documento>/seguimientos', methods=['GET'])
@login_required
def obtener_seguimientos(num_documento):
    """Obtener lista de seguimientos de un cliente"""
    return listar_seguimientos_cliente_api(num_documento)


@main_bp.route('/api/clientes/<num_documento>/seguimientos', methods=['POST'])
@login_required
def crear_seguimiento(num_documento):
    """Registrar un nuevo seguimiento para un cliente"""
    return registrar_seguimiento_api()


@main_bp.route('/api/clientes/<num_documento>/seguimientos/historial', methods=['GET'])
@login_required
def obtener_historial_seguimientos(num_documento):
    """Obtener historial completo de seguimientos de un cliente"""
    return listar_historial_seguimientos_api(num_documento)


@main_bp.route('/api/clientes/<num_documento>/seguimientos/ultimos-3', methods=['GET'])
@login_required
def obtener_ultimos_3_seguimientos(num_documento):
    """Obtener los últimos 3 seguimientos de un cliente"""
    return listar_ultimos_3_seguimientos_api(num_documento)


@main_bp.route('/api/tipos-seguimiento', methods=['GET'])
@login_required
def obtener_tipos_seguimiento():
    """Obtener lista de tipos de seguimiento disponibles"""
    return listar_tipos_seguimiento_api()


# ============================================================================
# RUTAS PARA ESTADÍSTICAS DE CLIENTES
# ============================================================================

@main_bp.route('/api/estadisticas/clientes-por-estado-prospeccion', methods=['GET'])
@login_required
def obtener_estadisticas_prospeccion():
    """Obtener estadísticas de clientes por estado de prospección"""
    return contar_clientes_por_estado_prospeccion_api()


@main_bp.route('/api/estadisticas/clientes-descartados', methods=['GET'])
@login_required
def obtener_estadisticas_descartados():
    """Obtener cantidad de clientes descartados"""
    return contar_clientes_descartados_api()


# ============================================================================
# RUTAS PARA CONFIGURACIÓN Y ADMINISTRACIÓN
# ============================================================================

@main_bp.route('/settings')
@login_required
def settings():
    """Página de configuración"""
    return render_template('settings.html')


@main_bp.route('/permisos-roles')
@login_required
def permisos_roles():
    """Página de administración de permisos y roles de usuarios"""
    return render_template('permissions.html')
