#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Funciones de gestion de clientes
- insertar_cliente_api()
- listar_clientes_api()
- listar_todos_clientes_api()
- eliminar_cliente_api()
"""

from flask import request
from mysql.connector import Error
from .funGeneral import get_db_connection


def insertar_cliente_api():
    """API para insertar un nuevo cliente"""
    try:
        from flask import session, current_app
        
        data = request.get_json()
        
        current_app.logger.info("=== INSERTAR NUEVO CLIENTE ===")
        # Convertir datos a string de forma segura para el log
        try:
            current_app.logger.info(f"Datos recibidos (total campos): {len(data)}")
        except:
            current_app.logger.info("Datos recibidos del formulario")
        
        # Obtener datos del formulario
        num_documento = data.get('num_documento', '').strip()
        tipo_documento = data.get('tipo_documento', 'DNI')
        nombres = data.get('nombres', '').strip()
        apellido_paterno = data.get('apellido_paterno', '').strip()
        apellido_materno = data.get('apellido_materno', '').strip()
        fecha_nacimiento = data.get('fecha_nacimiento', '')
        genero = data.get('genero', '')
        estado_civil = data.get('estado_civil', '')
        email = data.get('email', '').strip()
        celular = data.get('celular', '').strip()
        direccion = data.get('direccion', '').strip()
        id_distrito = data.get('id_distrito')
        
        # Obtener el documento del asesor logueado
        num_documento_asesor = session.get('user_documento', '')
        current_app.logger.info(f"Asesor logueado: {num_documento_asesor}")
        
        # Datos comerciales
        id_fuente_contacto = data.get('id_fuente_contacto')
        id_proyecto = data.get('id_proyecto')
        id_estado_prospeccion = data.get('id_estado_prospeccion')
        id_tipo_compra = data.get('id_tipo_compra')
        # Campo 'estado' eliminado - se usa estado_prospeccion
        # Campo 'fecha_proximo_seguimiento' eliminado - se maneja en TblClientesSeguimientos
        prioridad = data.get('prioridad', 'Media')
        observaciones = data.get('observaciones', '')
        
        # Validar campos requeridos
        if not all([num_documento, nombres, apellido_paterno, email, celular, direccion, id_distrito]):
            current_app.logger.warning(f"Campos requeridos faltantes al insertar cliente {num_documento}")
            return {'success': False, 'error': 'Por favor completa todos los campos requeridos'}, 400
        
        # Log seguro sin caracteres especiales que puedan causar problemas
        current_app.logger.info(f"Cliente a insertar - Doc: {num_documento}, Email: {email}")
        
        # Convertir valores vacios a None
        id_fuente_contacto = int(id_fuente_contacto) if id_fuente_contacto else None
        id_proyecto = int(id_proyecto) if id_proyecto else None
        id_estado_prospeccion = int(id_estado_prospeccion) if id_estado_prospeccion else None
        id_tipo_compra = int(id_tipo_compra) if id_tipo_compra else None
        id_distrito = int(id_distrito) if id_distrito else None
        
        connection = get_db_connection()
        if not connection:
            current_app.logger.error("Error de conexion a BD al insertar cliente")
            return {'success': False, 'error': 'Error de conexion'}, 500
        
        try:
            cursor = connection.cursor(dictionary=True)
            
            current_app.logger.info(f"Ejecutando SP sp_InsertarCliente para documento {num_documento}")
            
            # Llamar al SP (19 parámetros - SIN estado, SIN fecha_proximo_seguimiento)
            cursor.execute("""
                CALL sp_InsertarCliente(
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
                )
            """, (
                num_documento, tipo_documento, nombres, apellido_paterno, apellido_materno,
                fecha_nacimiento, genero, estado_civil,
                email, celular,
                direccion, id_distrito,
                num_documento_asesor,
                id_fuente_contacto, id_proyecto, id_estado_prospeccion, id_tipo_compra,
                prioridad, observaciones
            ))
            
            # El SP retorna SELECT con 'mensaje' y 'exito'
            result = cursor.fetchone()
            
            cursor.close()
            connection.close()
            
            if result and result.get('exito') == 1:
                current_app.logger.info(f"Cliente registrado exitosamente: {num_documento}")
                return {
                    'success': True,
                    'message': result.get('mensaje', 'Cliente registrado exitosamente')
                }, 201
            else:
                mensaje = result.get('mensaje', 'Error desconocido') if result else 'Error desconocido'
                current_app.logger.warning(f"No se pudo insertar cliente {num_documento}: {mensaje}")
                return {'success': False, 'error': mensaje}, 400
                
        except Error as e:
            current_app.logger.error(f"Error SQL al insertar cliente {num_documento}: {str(e)}")
            if connection.is_connected():
                connection.close()
            return {'success': False, 'error': f'Error en la base de datos: {str(e)}'}, 500
    
    except Exception as e:
        from flask import current_app
        current_app.logger.error(f"Error general al insertar cliente", exc_info=True)
        return {'success': False, 'error': f'Error del servidor: {str(e)}'}, 500


def listar_clientes_api():
    """API para listar clientes del asesor logueado (o todos si es ADMINISTRADOR)"""
    try:
        from flask import session, current_app
        from datetime import datetime
        
        # Obtener el documento y rol del asesor logueado
        num_documento_asesor = session.get('user_documento', '')
        rol_usuario = session.get('user_role', '').upper()  # Convertir a mayusculas
        
        # Log de inicio
        current_app.logger.info(f"=== LISTAR CLIENTES (SIN FILTROS) ===")
        current_app.logger.info(f"Usuario: {num_documento_asesor}")
        current_app.logger.info(f"Rol original: {session.get('user_role', '')}")
        current_app.logger.info(f"Rol convertido: {rol_usuario}")
        
        if not num_documento_asesor:
            current_app.logger.warning("Usuario no autenticado intentando listar clientes")
            return {'success': False, 'error': 'Usuario no autenticado'}, 401
        
        connection = get_db_connection()
        if not connection:
            current_app.logger.error("Error al conectar a la base de datos")
            return {'success': False, 'error': 'Error de conexion'}, 500
        
        try:
            cursor = connection.cursor(dictionary=True)
            
            # Si es ADMINISTRADOR, mostrar TODOS los clientes SIN FILTROS
            if rol_usuario == 'ADMINISTRADOR':
                current_app.logger.info("Usuario ADMINISTRADOR - Ejecutando sp_ListarTodosLosClientes() SIN FILTROS")
                # Llamar SP con parámetros NULL para obtener todos los clientes
                cursor.execute("""
                    CALL sp_ListarTodosLosClientes(
                        NULL,  -- p_nombre
                        NULL,  -- p_estado_sys
                        NULL,  -- p_id_estado_prospeccion
                        NULL,  -- p_fecha_desde
                        NULL   -- p_fecha_hasta
                    )
                """)
            else:
                # Si es asesor normal, solo mostrar sus clientes
                current_app.logger.info(f"Usuario {rol_usuario} - Ejecutando sp_ListarClientes({num_documento_asesor}) SIN FILTROS")
                cursor.execute("""
                    CALL sp_ListarClientes(
                        %s,    -- p_num_documento_asesor
                        NULL,  -- p_nombre
                        NULL,  -- p_estado_sys
                        NULL,  -- p_id_estado_prospeccion
                        NULL,  -- p_fecha_desde
                        NULL   -- p_fecha_hasta
                    )
                """, (num_documento_asesor,))
            
            clientes = cursor.fetchall()
            
            # Convertir datetime a string para JSON
            for cliente in clientes:
                if cliente.get('fecha_proximo_seguimiento') and isinstance(cliente['fecha_proximo_seguimiento'], datetime):
                    cliente['fecha_proximo_seguimiento'] = cliente['fecha_proximo_seguimiento'].isoformat()
                if cliente.get('fecha_nacimiento') and isinstance(cliente['fecha_nacimiento'], datetime):
                    cliente['fecha_nacimiento'] = cliente['fecha_nacimiento'].isoformat()
                if cliente.get('fecha_creacion') and isinstance(cliente['fecha_creacion'], datetime):
                    cliente['fecha_creacion'] = cliente['fecha_creacion'].isoformat()
                if cliente.get('fecha_actualizacion') and isinstance(cliente['fecha_actualizacion'], datetime):
                    cliente['fecha_actualizacion'] = cliente['fecha_actualizacion'].isoformat()
                if cliente.get('ultimo_seguimiento_fecha') and isinstance(cliente['ultimo_seguimiento_fecha'], datetime):
                    cliente['ultimo_seguimiento_fecha'] = cliente['ultimo_seguimiento_fecha'].isoformat()
            
            current_app.logger.info(f"Se obtuvieron {len(clientes)} clientes")
            
            cursor.close()
            connection.close()
            
            return {
                'success': True,
                'data': clientes,
                'total': len(clientes),
                'es_admin': rol_usuario == 'ADMINISTRADOR'
            }, 200
            
        except Error as e:
            current_app.logger.error(f"Error SQL: {e}")
            if connection.is_connected():
                connection.close()
            return {'success': False, 'error': f'Error en la base de datos: {str(e)}'}, 500
    
    except Exception as e:
        current_app.logger.error(f"Error general: {e}", exc_info=True)
        return {'success': False, 'error': f'Error del servidor: {str(e)}'}, 500


def listar_todos_clientes_api():
    """API para listar TODOS los clientes (para administradores) - DEPRECATED"""
    try:
        from flask import current_app
        
        current_app.logger.info("=== LISTAR TODOS LOS CLIENTES (DEPRECATED) ===")
        
        # Esta función está deprecada, usar filtrar_clientes_api en su lugar
        return {'success': False, 'error': 'Use /api/clientes con filtros en su lugar'}, 400
    
    except Exception as e:
        from flask import current_app
        current_app.logger.error(f"Error: {e}", exc_info=True)
        return {'success': False, 'error': f'Error del servidor: {str(e)}'}, 500


def eliminar_cliente_api(num_documento):
    """API para eliminar un cliente por numero de documento"""
    try:
        from flask import session, current_app
        
        # Verificar que el usuario este autenticado
        if not session.get('user_documento'):
            current_app.logger.warning(f"Intento de eliminar cliente sin autenticacion")
            return {'success': False, 'error': 'Usuario no autenticado'}, 401
        
        connection = get_db_connection()
        if not connection:
            current_app.logger.error("Error de conexion al eliminar cliente")
            return {'success': False, 'error': 'Error de conexion'}, 500
        
        try:
            cursor = connection.cursor()
            
            # Variables para los parametros OUT
            cursor.execute("SET @p_eliminado = 0")
            cursor.execute("SET @p_mensaje = ''")
            
            # Llamar al SP para eliminar cliente
            current_app.logger.info(f"Intentando eliminar cliente con documento: {num_documento}")
            cursor.execute("CALL sp_EliminarCliente(%s, @p_eliminado, @p_mensaje)", (num_documento,))
            
            # Obtener los valores de salida
            cursor.execute("SELECT @p_eliminado AS eliminado, @p_mensaje AS mensaje")
            result = cursor.fetchone()
            
            cursor.close()
            connection.close()
            
            if result[0] == 1:  # p_eliminado == 1
                current_app.logger.info(f"Cliente eliminado exitosamente: {num_documento}")
                return {
                    'success': True,
                    'message': result[1]  # p_mensaje
                }, 200
            else:
                current_app.logger.warning(f"No se pudo eliminar cliente {num_documento}: {result[1]}")
                return {
                    'success': False,
                    'error': result[1]
                }, 400
            
        except Error as e:
            current_app.logger.error(f"Error SQL al eliminar cliente {num_documento}: {e}")
            if connection.is_connected():
                connection.close()
            return {'success': False, 'error': f'Error en la base de datos: {str(e)}'}, 500
    
    except Exception as e:
        from flask import current_app
        current_app.logger.error(f"Error general al eliminar cliente: {e}", exc_info=True)
        return {'success': False, 'error': f'Error del servidor: {str(e)}'}, 500


def obtener_cliente_por_documento_api(num_documento):
    """API para obtener los datos completos de un cliente por su número de documento"""
    try:
        from flask import current_app
        
        current_app.logger.info(f"=== OBTENER CLIENTE POR DOCUMENTO ===")
        current_app.logger.info(f"Documento: {num_documento}")
        
        connection = get_db_connection()
        if not connection:
            current_app.logger.error("Error de conexión al obtener cliente")
            return {'success': False, 'error': 'Error de conexión'}, 500
        
        try:
            cursor = connection.cursor(dictionary=True)
            
            # Llamar al SP para obtener los datos del cliente
            current_app.logger.info(f"Ejecutando SP sp_ObtenerClientePorDocumento({num_documento})")
            cursor.execute("CALL sp_ObtenerClientePorDocumento(%s)", (num_documento,))
            
            cliente = cursor.fetchone()
            
            cursor.close()
            connection.close()
            
            if cliente:
                current_app.logger.info(f"Cliente encontrado: {num_documento}")
                return {
                    'success': True,
                    'data': cliente
                }, 200
            else:
                current_app.logger.warning(f"Cliente no encontrado: {num_documento}")
                return {
                    'success': False,
                    'error': 'Cliente no encontrado'
                }, 404
            
        except Error as e:
            current_app.logger.error(f"Error SQL al obtener cliente {num_documento}: {e}")
            if connection.is_connected():
                connection.close()
            return {'success': False, 'error': f'Error en la base de datos: {str(e)}'}, 500
    
    except Exception as e:
        from flask import current_app
        current_app.logger.error(f"Error general al obtener cliente: {e}", exc_info=True)
        return {'success': False, 'error': f'Error del servidor: {str(e)}'}, 500


def actualizar_cliente_api(num_documento):
    """API para actualizar los datos de un cliente existente"""
    try:
        from flask import session, current_app
        
        data = request.get_json()
        
        current_app.logger.info("=== ACTUALIZAR CLIENTE ===")
        current_app.logger.info(f"Documento a actualizar: {num_documento}")
        
        # Obtener datos del formulario
        tipo_documento = data.get('tipo_documento', 'DNI')
        nombres = data.get('nombres', '').strip()
        apellido_paterno = data.get('apellido_paterno', '').strip()
        apellido_materno = data.get('apellido_materno', '').strip()
        fecha_nacimiento = data.get('fecha_nacimiento', '')
        genero = data.get('genero', '')
        estado_civil = data.get('estado_civil', '')
        email = data.get('email', '').strip()
        celular = data.get('celular', '').strip()
        direccion = data.get('direccion', '').strip()
        id_distrito = data.get('id_distrito')
        
        # Obtener el documento del asesor logueado (quien modifica)
        num_documento_modificador = session.get('user_documento', '')
        current_app.logger.info(f"Usuario que modifica: {num_documento_modificador}")
        
        # Datos comerciales
        id_fuente_contacto = data.get('id_fuente_contacto')
        id_proyecto = data.get('id_proyecto')
        id_estado_prospeccion = data.get('id_estado_prospeccion')
        id_tipo_compra = data.get('id_tipo_compra')
        # Campo 'estado' eliminado - se usa estado_prospeccion
        # Campo 'fecha_proximo_seguimiento' eliminado - se maneja en TblClientesSeguimientos
        prioridad = data.get('prioridad', 'Media')
        observaciones = data.get('observaciones', '')
        
        # Validar campos requeridos
        if not all([num_documento, nombres, apellido_paterno, email, celular, direccion, id_distrito]):
            current_app.logger.warning(f"Campos requeridos faltantes al actualizar cliente {num_documento}")
            return {'success': False, 'error': 'Por favor completa todos los campos requeridos'}, 400
        
        current_app.logger.info(f"Cliente a actualizar - Doc: {num_documento}, Email: {email}")
        
        # Convertir valores vacíos a None
        id_fuente_contacto = int(id_fuente_contacto) if id_fuente_contacto else None
        id_proyecto = int(id_proyecto) if id_proyecto else None
        id_estado_prospeccion = int(id_estado_prospeccion) if id_estado_prospeccion else None
        id_tipo_compra = int(id_tipo_compra) if id_tipo_compra else None
        id_distrito = int(id_distrito) if id_distrito else None
        
        connection = get_db_connection()
        if not connection:
            current_app.logger.error("Error de conexión a BD al actualizar cliente")
            return {'success': False, 'error': 'Error de conexión'}, 500
        
        try:
            cursor = connection.cursor(dictionary=True)
            
            current_app.logger.info(f"Ejecutando SP sp_ActualizarCliente para documento {num_documento}")
            
            # Llamar al SP (18 parámetros - SIN estado, SIN fecha_proximo_seguimiento)
            cursor.execute("""
                CALL sp_ActualizarCliente(
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
                )
            """, (
                num_documento, tipo_documento, nombres, apellido_paterno, apellido_materno,
                fecha_nacimiento, genero, estado_civil,
                email, celular,
                direccion, id_distrito,
                id_fuente_contacto, id_proyecto, id_estado_prospeccion, id_tipo_compra,
                prioridad, observaciones
            ))
            
            # El SP retorna SELECT con 'mensaje' y 'exito'
            result = cursor.fetchone()
            
            cursor.close()
            connection.close()
            
            if result and result.get('exito') == 1:
                current_app.logger.info(f"Cliente actualizado exitosamente: {num_documento}")
                return {
                    'success': True,
                    'message': result.get('mensaje', 'Cliente actualizado exitosamente')
                }, 200
            else:
                mensaje = result.get('mensaje', 'Error desconocido') if result else 'Error desconocido'
                current_app.logger.warning(f"No se pudo actualizar cliente {num_documento}: {mensaje}")
                return {'success': False, 'error': mensaje}, 400
                
        except Error as e:
            current_app.logger.error(f"Error SQL al actualizar cliente {num_documento}: {str(e)}")
            if connection.is_connected():
                connection.close()
            return {'success': False, 'error': f'Error en la base de datos: {str(e)}'}, 500
    
    except Exception as e:
        from flask import current_app
        current_app.logger.error(f"Error general al actualizar cliente", exc_info=True)
        return {'success': False, 'error': f'Error del servidor: {str(e)}'}, 500


def descartar_cliente_api(num_documento):
    """API para alternar estado del cliente entre ACTIVO y DESCARTADO usando SP"""
    try:
        from flask import current_app
        
        current_app.logger.info("=== ALTERNAR ESTADO CLIENTE (SP) ===")
        current_app.logger.info(f"Documento: {num_documento}")
        
        connection = get_db_connection()
        if not connection:
            current_app.logger.error("Error de conexión")
            return {'success': False, 'error': 'Error de conexión'}, 500
        
        try:
            cursor = connection.cursor(dictionary=True)
            
            # Ejecutar SP directamente con execute (mejor compatibilidad)
            query = f"CALL sp_AlternarEstadoCliente('{num_documento}')"
            current_app.logger.info(f"Query: {query}")
            
            cursor.execute(query)
            
            # Obtener resultado del SELECT del SP
            resultado = cursor.fetchone()
            current_app.logger.info(f"Resultado fetchone: {resultado}")
            
            # Si fetchone no funciona, intentar fetchall
            if not resultado:
                cursor.reset()
                cursor.execute(query)
                resultados = cursor.fetchall()
                current_app.logger.info(f"Resultados fetchall: {resultados}")
                
                if resultados and len(resultados) > 0:
                    resultado = resultados[0]
            
            if not resultado:
                current_app.logger.error(f"SP retornó resultado vacío")
                cursor.close()
                connection.close()
                return {'success': False, 'error': 'Error al ejecutar procedimiento'}, 500
            
            estado_resultado = resultado.get('resultado')
            estado_anterior = resultado.get('estado_anterior')
            estado_nuevo = resultado.get('estado_nuevo')
            
            current_app.logger.info(f"Resultado SP: {estado_resultado}")
            current_app.logger.info(f"Estado: {estado_anterior} → {estado_nuevo}")
            
            # Consumir todos los result sets
            try:
                while cursor.nextset():
                    pass
            except:
                pass
            
            if estado_resultado == 'OK':
                connection.commit()
                current_app.logger.info(f"Estado alterado a {estado_nuevo}")
                cursor.close()
                connection.close()
                
                return {
                    'success': True,
                    'message': f'Cliente {estado_nuevo.lower()} exitosamente',
                    'nuevo_estado': estado_nuevo,
                    'estado_anterior': estado_anterior
                }, 200
            
            elif estado_resultado == 'CLIENTE_NO_ENCONTRADO':
                current_app.logger.warning(f"Cliente no encontrado: {num_documento}")
                cursor.close()
                connection.close()
                return {'success': False, 'error': 'Cliente no encontrado'}, 404
            
            else:
                current_app.logger.error(f"Error en SP: {estado_resultado}")
                cursor.close()
                connection.close()
                return {'success': False, 'error': f'Error en procedimiento: {estado_resultado}'}, 400
                
        except Error as e:
            current_app.logger.error(f"Error SQL: {str(e)}")
            if connection.is_connected():
                connection.rollback()
                connection.close()
            return {'success': False, 'error': f'Error en la base de datos: {str(e)}'}, 500
        except Exception as inner_e:
            current_app.logger.error(f"Error: {str(inner_e)}", exc_info=True)
            if connection.is_connected():
                connection.rollback()
                connection.close()
            return {'success': False, 'error': f'Error: {str(inner_e)}'}, 500
    
    except Exception as e:
        from flask import current_app
        current_app.logger.error(f"Error general", exc_info=True)
        return {'success': False, 'error': f'Error del servidor: {str(e)}'}, 500


# =====================================================
# FUNCIONES DE SEGUIMIENTO DE CLIENTES
# =====================================================

def listar_seguimientos_cliente_api(num_documento):
    """API para listar seguimientos de un cliente"""
    try:
        from flask import current_app
        
        current_app.logger.info(f"=== LISTAR SEGUIMIENTOS cliente: {num_documento} ===")
        
        connection = get_db_connection()
        if not connection:
            current_app.logger.error("Error de conexión")
            return {'success': False, 'error': 'Error de conexión'}, 500
        
        try:
            cursor = connection.cursor(dictionary=True)
            
            # Llamar al SP
            cursor.execute("CALL sp_ListarSeguimientosCliente(%s)", (num_documento,))
            seguimientos = cursor.fetchall()
            
            cursor.close()
            connection.close()
            
            current_app.logger.info(f"Se obtuvieron {len(seguimientos)} seguimientos")
            
            return {
                'success': True,
                'data': seguimientos,
                'total': len(seguimientos)
            }, 200
            
        except Error as e:
            current_app.logger.error(f"Error SQL: {e}")
            if connection.is_connected():
                connection.close()
            return {'success': False, 'error': f'Error en la base de datos: {str(e)}'}, 500
    
    except Exception as e:
        from flask import current_app
        current_app.logger.error(f"Error general: {e}", exc_info=True)
        return {'success': False, 'error': 'Error del servidor'}, 500


def listar_tipos_seguimiento_api():
    """API para listar tipos de seguimiento disponibles"""
    try:
        from flask import current_app
        
        current_app.logger.info("=== LISTAR TIPOS DE SEGUIMIENTO ===")
        
        connection = get_db_connection()
        if not connection:
            return {'success': False, 'error': 'Error de conexión'}, 500
        
        try:
            cursor = connection.cursor(dictionary=True)
            
            # Llamar al SP
            cursor.execute("CALL sp_ListarTiposSeguimiento()")
            tipos = cursor.fetchall()
            
            cursor.close()
            connection.close()
            
            current_app.logger.info(f"Se obtuvieron {len(tipos)} tipos de seguimiento")
            
            return {
                'success': True,
                'data': tipos,
                'total': len(tipos)
            }, 200
            
        except Error as e:
            current_app.logger.error(f"Error SQL: {e}")
            if connection.is_connected():
                connection.close()
            return {'success': False, 'error': f'Error en la base de datos: {str(e)}'}, 500
    
    except Exception as e:
        from flask import current_app
        current_app.logger.error(f"Error general: {e}", exc_info=True)
        return {'success': False, 'error': 'Error del servidor'}, 500


def listar_historial_seguimientos_api(num_documento):
    """API para listar el historial completo de seguimientos de un cliente"""
    try:
        from flask import current_app
        from datetime import datetime
        
        current_app.logger.info(f"=== LISTAR HISTORIAL SEGUIMIENTOS ===")
        current_app.logger.info(f"Cliente: {num_documento}")
        
        connection = get_db_connection()
        if not connection:
            current_app.logger.error("Error de conexión a BD")
            return {'success': True, 'data': [], 'total': 0}, 200
        
        try:
            cursor = connection.cursor(dictionary=True)
            
            # Llamar al SP directamente
            current_app.logger.info(f"Ejecutando SP sp_ListarHistorialSeguimientos({num_documento})")
            cursor.execute("CALL sp_ListarHistorialSeguimientos(%s)", (num_documento,))
            
            historial = cursor.fetchall()
            
            current_app.logger.info(f"Se obtuvieron {len(historial)} seguimientos")
            
            # Convertir datetime a string para JSON
            for seg in historial:
                if seg.get('fecha_seguimiento') and isinstance(seg['fecha_seguimiento'], datetime):
                    seg['fecha_seguimiento'] = seg['fecha_seguimiento'].isoformat()
                if seg.get('fecha_creacion') and isinstance(seg['fecha_creacion'], datetime):
                    seg['fecha_creacion'] = seg['fecha_creacion'].isoformat()
            
            cursor.close()
            connection.close()
            
            current_app.logger.info(f"Historial cargado correctamente")
            
            return {
                'success': True,
                'data': historial,
                'total': len(historial)
            }, 200
            
        except Exception as e:
            current_app.logger.error(f"Error SQL al obtener historial: {str(e)}", exc_info=True)
            if connection.is_connected():
                connection.close()
            return {'success': True, 'data': [], 'total': 0}, 200
    
    except Exception as e:
        from flask import current_app
        current_app.logger.error(f"Error general al obtener historial: {str(e)}", exc_info=True)
        return {'success': True, 'data': [], 'total': 0}, 200


# =====================================================
# FUNCIÓN PARA LISTAR ÚLTIMOS 3 SEGUIMIENTOS
# =====================================================

def listar_ultimos_3_seguimientos_api(num_documento):
    """API para listar los últimos 3 seguimientos de un cliente"""
    try:
        from flask import current_app
        from datetime import datetime
        
        current_app.logger.info(f"=== LISTAR ÚLTIMOS 3 SEGUIMIENTOS ===")
        current_app.logger.info(f"Cliente: {num_documento}")
        
        connection = get_db_connection()
        if not connection:
            current_app.logger.error("Error de conexión a BD")
            return {'success': True, 'data': [], 'total': 0}, 200
        
        try:
            cursor = connection.cursor(dictionary=True)
            
            # Llamar al SP directamente
            current_app.logger.info(f"Ejecutando SP sp_Ultimos3Seguimientos({num_documento})")
            cursor.execute("CALL sp_Ultimos3Seguimientos(%s)", (num_documento,))
            
            ultimos3 = cursor.fetchall()
            
            current_app.logger.info(f"Se obtuvieron {len(ultimos3)} últimos seguimientos")
            
            # Convertir datetime a string para JSON
            for seg in ultimos3:
                if seg.get('fecha_seguimiento') and isinstance(seg['fecha_seguimiento'], datetime):
                    seg['fecha_seguimiento'] = seg['fecha_seguimiento'].isoformat()
                if seg.get('fecha_registro') and isinstance(seg['fecha_registro'], datetime):
                    seg['fecha_registro'] = seg['fecha_registro'].isoformat()
            
            cursor.close()
            connection.close()
            
            current_app.logger.info(f"Últimos 3 seguimientos cargados correctamente")
            
            return {
                'success': True,
                'data': ultimos3,
                'total': len(ultimos3)
            }, 200
            
        except Exception as e:
            current_app.logger.error(f"Error SQL al obtener últimos 3 seguimientos: {str(e)}", exc_info=True)
            if connection.is_connected():
                connection.close()
            return {'success': True, 'data': [], 'total': 0}, 200
    
    except Exception as e:
        from flask import current_app
        current_app.logger.error(f"Error general al obtener últimos 3 seguimientos: {str(e)}", exc_info=True)
        return {'success': True, 'data': [], 'total': 0}, 200


def contar_clientes_por_estado_prospeccion_api():
    """API para obtener cantidad de clientes por estado de prospección"""
    try:
        from flask import current_app
        
        connection = get_db_connection()
        if not connection:
            return {'success': False, 'error': 'Error de conexión'}, 500
        
        try:
            cursor = connection.cursor(dictionary=True)
            
            # Llamar al SP
            cursor.execute("CALL sp_ContarClientesPorEstadoProspeccion()")
            
            estadisticas = cursor.fetchall()
            
            cursor.close()
            connection.close()
            
            if estadisticas:
                return {
                    'success': True,
                    'data': estadisticas
                }, 200
            else:
                return {
                    'success': True,
                    'data': []
                }, 200
            
        except Error as e:
            current_app.logger.error(f"Error SQL: {str(e)}")
            if connection.is_connected():
                connection.close()
            return {'success': False, 'error': f'Error en BD: {str(e)}'}, 500
    
    except Exception as e:
        from flask import current_app
        current_app.logger.error(f"Error general: {str(e)}", exc_info=True)
        return {'success': False, 'error': f'Error: {str(e)}'}, 500


def contar_clientes_descartados_api():
    """API para obtener cantidad de clientes descartados (por asesor o total si es admin)"""
    try:
        from flask import current_app, session
        
        num_documento_asesor = session.get('user_documento', '')
        rol_usuario = session.get('user_role', '').upper()
        
        if not num_documento_asesor:
            return {'success': False, 'error': 'Usuario no autenticado'}, 401
        
        connection = get_db_connection()
        if not connection:
            return {'success': False, 'error': 'Error de conexión'}, 500
        
        try:
            cursor = connection.cursor(dictionary=True)
            
            # Si es admin, obtener todos los descartados
            if rol_usuario == 'ADMINISTRADOR':
                cursor.execute("CALL sp_ContarClientesDescartadosAdmin()")
            else:
                # Si es asesor, obtener solo sus descartados
                cursor.execute("CALL sp_ContarClientesDescartadosAsesor(%s)", (num_documento_asesor,))
            
            result = cursor.fetchone()
            cantidad = result.get('cantidad_descartados', 0) if result else 0
            
            cursor.close()
            connection.close()
            
            return {
                'success': True,
                'data': cantidad
            }, 200
            
        except Error as e:
            current_app.logger.error(f"Error SQL: {str(e)}")
            if connection.is_connected():
                connection.close()
            return {'success': False, 'error': f'Error en BD: {str(e)}'}, 500
    
    except Exception as e:
        from flask import current_app
        current_app.logger.error(f"Error general: {str(e)}", exc_info=True)
        return {'success': False, 'error': f'Error: {str(e)}'}, 500


# =====================================================
# FUNCIÓN PARA REGISTRAR SEGUIMIENTO
# =====================================================

def registrar_seguimiento_api():
    """API para registrar un nuevo seguimiento de cliente según sp_RegistrarSeguimiento"""
    try:
        from flask import session, current_app
        
        data = request.get_json()
        
        current_app.logger.info("=== REGISTRAR SEGUIMIENTO ===")
        
        # Obtener datos del formulario
        num_documento_cliente = data.get('num_documento', '').strip()
        id_tipo_seguimiento = data.get('id_tipo_seguimiento')
        fecha_seguimiento = data.get('fecha_seguimiento', '')
        observacion = data.get('observacion', '').strip()
        
        # Obtener el documento del asesor logueado
        realizado_por = session.get('user_documento', '')
        
        current_app.logger.info(f"Cliente: {num_documento_cliente}, Tipo: {id_tipo_seguimiento}, Asesor: {realizado_por}")
        
        # Validar campos requeridos
        if not all([num_documento_cliente, id_tipo_seguimiento, fecha_seguimiento, observacion, realizado_por]):
            current_app.logger.warning("Campos requeridos faltantes al registrar seguimiento")
            return {'success': False, 'error': 'Por favor completa todos los campos'}, 400
        
        # Convertir id_tipo_seguimiento a int
        try:
            id_tipo_seguimiento = int(id_tipo_seguimiento)
        except (ValueError, TypeError):
            current_app.logger.error(f"id_tipo_seguimiento inválido: {id_tipo_seguimiento}")
            return {'success': False, 'error': 'Tipo de seguimiento inválido'}, 400
        
        connection = get_db_connection()
        if not connection:
            current_app.logger.error("Error de conexión a BD al registrar seguimiento")
            return {'success': False, 'error': 'Error de conexión'}, 500
        
        try:
            cursor = connection.cursor(dictionary=True)
            
            current_app.logger.info(f"Ejecutando SP sp_RegistrarSeguimiento")
            
            # El SP espera parámetros IN y OUT
            # Inicializar variables OUT
            cursor.execute("SET @p_id_seguimiento = 0")
            cursor.execute("SET @p_mensaje = ''")
            
            # Llamar al SP con los parámetros correctos
            cursor.execute("""
                CALL sp_RegistrarSeguimiento(
                    %s, %s, %s, %s, %s, @p_id_seguimiento, @p_mensaje
                )
            """, (
                num_documento_cliente,
                id_tipo_seguimiento,
                fecha_seguimiento,
                observacion,
                realizado_por
            ))
            
            # Obtener los valores de salida
            cursor.execute("SELECT @p_id_seguimiento AS id_seguimiento, @p_mensaje AS mensaje")
            result = cursor.fetchone()
            
            cursor.close()
            connection.close()
            
            if result:
                id_seguimiento = result.get('id_seguimiento')
                mensaje = result.get('mensaje', 'Error desconocido')
                
                # El SP retorna un id > 0 si fue exitoso
                if id_seguimiento and id_seguimiento > 0:
                    current_app.logger.info(f"Seguimiento registrado exitosamente: ID {id_seguimiento}")
                    return {
                        'success': True,
                        'message': 'Seguimiento registrado exitosamente',
                        'id_seguimiento': id_seguimiento
                    }, 201
                else:
                    current_app.logger.warning(f"Error al registrar seguimiento: {mensaje}")
                    return {'success': False, 'error': mensaje}, 400
            else:
                current_app.logger.error("No se recibieron datos de salida del SP")
                return {'success': False, 'error': 'Error al registrar el seguimiento'}, 500
                
        except Error as e:
            current_app.logger.error(f"Error SQL al registrar seguimiento: {str(e)}")
            if connection.is_connected():
                connection.close()
            return {'success': False, 'error': f'Error en la base de datos: {str(e)}'}, 500
    
    except Exception as e:
        from flask import current_app
        current_app.logger.error(f"Error general al registrar seguimiento", exc_info=True)
        return {'success': False, 'error': f'Error del servidor: {str(e)}'}, 500


# =====================================================
# FUNCIÓN PARA FILTRAR CLIENTES
# =====================================================

def filtrar_clientes_api():
    """API para filtrar clientes - Usa SPs con parámetros de filtro"""
    try:
        from flask import session, current_app
        from datetime import datetime
        
        current_app.logger.info("=== FILTRAR CLIENTES (API) ===")
        
        # Obtener parámetros de query
        nombre_busqueda = request.args.get('nombre', '').strip() if request.args.get('nombre') else None
        estado_sys = request.args.get('estado', '').strip().upper() if request.args.get('estado') else None
        id_estado_prospeccion_str = request.args.get('estado_prospeccion', '').strip()
        fecha_desde_str = request.args.get('fecha_desde', '').strip() if request.args.get('fecha_desde') else None
        fecha_hasta_str = request.args.get('fecha_hasta', '').strip() if request.args.get('fecha_hasta') else None
        
        current_app.logger.info(f"Parámetros recibidos:")
        current_app.logger.info(f"  nombre: '{nombre_busqueda}'")
        current_app.logger.info(f"  estado: '{estado_sys}'")
        current_app.logger.info(f"  estado_prospeccion: '{id_estado_prospeccion_str}'")
        current_app.logger.info(f"  fecha_desde: '{fecha_desde_str}'")
        current_app.logger.info(f"  fecha_hasta: '{fecha_hasta_str}'")
        
        # Convertir id_estado_prospeccion a int o None
        id_estado_prospeccion = None
        if id_estado_prospeccion_str:
            try:
                id_estado_prospeccion = int(id_estado_prospeccion_str)
                current_app.logger.info(f"  id_estado_prospeccion convertido a: {id_estado_prospeccion}")
            except ValueError:
                current_app.logger.warning(f"  No se pudo convertir id_estado_prospeccion: {id_estado_prospeccion_str}")
                id_estado_prospeccion = None
        
        # Obtener el documento y rol del asesor logueado
        num_documento_asesor = session.get('user_documento', '')
        rol_usuario = session.get('user_role', '').upper()
        
        current_app.logger.info(f"Usuario: {num_documento_asesor}, Rol: {rol_usuario}")
        
        if not num_documento_asesor:
            current_app.logger.warning("Usuario no autenticado intentando filtrar clientes")
            return {'success': False, 'error': 'Usuario no autenticado'}, 401
        
        # Convertir strings de fecha a objetos date si son válidos
        fecha_desde = None
        fecha_hasta = None
        
        try:
            if fecha_desde_str:
                fecha_desde = datetime.strptime(fecha_desde_str, '%Y-%m-%d').date()
                current_app.logger.info(f"  fecha_desde convertida a: {fecha_desde}")
            if fecha_hasta_str:
                fecha_hasta = datetime.strptime(fecha_hasta_str, '%Y-%m-%d').date()
                current_app.logger.info(f"  fecha_hasta convertida a: {fecha_hasta}")
        except ValueError as e:
            current_app.logger.error(f"Error al parsear fechas: {e}")
            return {'success': False, 'error': 'Formato de fecha inválido'}, 400
        
        connection = get_db_connection()
        if not connection:
            current_app.logger.error("Error al conectar a la base de datos")
            return {'success': False, 'error': 'Error de conexión'}, 500
        
        try:
            cursor = connection.cursor(dictionary=True)
            
            # Si es ADMINISTRADOR, usar sp_ListarTodosLosClientes CON FILTROS
            if rol_usuario == 'ADMINISTRADOR':
                current_app.logger.info("Ejecutando sp_ListarTodosLosClientes() con filtros para ADMIN")
                cursor.execute("""
                    CALL sp_ListarTodosLosClientes(
                        %s,  -- p_nombre
                        %s,  -- p_estado_sys
                        %s,  -- p_id_estado_prospeccion
                        %s,  -- p_fecha_desde
                        %s   -- p_fecha_hasta
                    )
                """, (
                    nombre_busqueda,
                    estado_sys,
                    id_estado_prospeccion,
                    fecha_desde,
                    fecha_hasta
                ))
            else:
                # Si es asesor normal, usar sp_ListarClientes CON FILTROS
                current_app.logger.info(f"Ejecutando sp_ListarClientes() con filtros para ASESOR: {num_documento_asesor}")
                cursor.execute("""
                    CALL sp_ListarClientes(
                        %s,  -- p_num_documento_asesor
                        %s,  -- p_nombre
                        %s,  -- p_estado_sys
                        %s,  -- p_id_estado_prospeccion
                        %s,  -- p_fecha_desde
                        %s   -- p_fecha_hasta
                    )
                """, (
                    num_documento_asesor,
                    nombre_busqueda,
                    estado_sys,
                    id_estado_prospeccion,
                    fecha_desde,
                    fecha_hasta
                ))
            
            clientes = cursor.fetchall()
            current_app.logger.info(f"SP retornó {len(clientes)} clientes")
            
            # Convertir datetime a string para JSON
            for cliente in clientes:
                if cliente.get('fecha_proximo_seguimiento') and isinstance(cliente['fecha_proximo_seguimiento'], datetime):
                    cliente['fecha_proximo_seguimiento'] = cliente['fecha_proximo_seguimiento'].isoformat()
                if cliente.get('fecha_nacimiento') and isinstance(cliente['fecha_nacimiento'], datetime):
                    cliente['fecha_nacimiento'] = cliente['fecha_nacimiento'].isoformat()
                if cliente.get('fecha_creacion') and isinstance(cliente['fecha_creacion'], datetime):
                    cliente['fecha_creacion'] = cliente['fecha_creacion'].isoformat()
                if cliente.get('fecha_actualizacion') and isinstance(cliente['fecha_actualizacion'], datetime):
                    cliente['fecha_actualizacion'] = cliente['fecha_actualizacion'].isoformat()
                if cliente.get('ultimo_seguimiento_fecha') and isinstance(cliente['ultimo_seguimiento_fecha'], datetime):
                    cliente['ultimo_seguimiento_fecha'] = cliente['ultimo_seguimiento_fecha'].isoformat()
            
            cursor.close()
            connection.close()
            
            current_app.logger.info(f"Filtrado completado. Retornando {len(clientes)} clientes")
            
            return {
                'success': True,
                'data': clientes,
                'total': len(clientes)
            }, 200
            
        except Error as e:
            current_app.logger.error(f"Error SQL al filtrar clientes: {e}")
            print(f"Error SQL: {e}")
            if connection.is_connected():
                connection.close()
            return {'success': False, 'error': f'Error en la base de datos: {str(e)}'}, 500
    
    except Exception as e:
        from flask import current_app
        current_app.logger.error(f"Error general al filtrar clientes: {e}", exc_info=True)
        print(f"Error general: {e}")
        import traceback
        traceback.print_exc()
        return {'success': False, 'error': f'Error del servidor: {str(e)}'}, 500


# =====================================================
# IMPORTACION MASIVA DE CLIENTES
# =====================================================

def importar_clientes_api():
    """Importar clientes desde archivo Excel - Usando SP"""
    try:
        from flask import session, request, jsonify, current_app
        from openpyxl import load_workbook
        import logging
        
        logger = logging.getLogger(__name__)
        
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': 'No se envió archivo'}), 400
        
        archivo = request.files['file']
        if archivo.filename == '':
            return jsonify({'success': False, 'message': 'No se seleccionó archivo'}), 400
        
        if not archivo.filename.endswith('.xlsx'):
            return jsonify({'success': False, 'message': 'El archivo debe ser Excel (.xlsx)'}), 400
        
        wb = load_workbook(archivo)
        ws = wb.active
        
        clientes_importados = 0
        errores = []
        num_documento_asesor = session.get('user_documento', '999999999')
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                tipo_documento = row[0]
                num_documento = str(row[1]) if row[1] else None
                genero = row[2]
                nombres = row[3]
                apellido_paterno = row[4]
                apellido_materno = row[5] if row[5] else None
                fecha_nacimiento = row[6]
                estado_civil = row[7]
                email = row[8]
                celular = row[9]
                direccion = row[10]
                departamento_nombre = row[11]
                provincia_nombre = row[12]
                distrito_nombre = row[13]
                fuente_contacto = row[14] if row[14] else None
                proyecto_interes = row[15] if row[15] else None
                estado_prospeccion = row[16] if row[16] else None
                tipo_compra = row[17] if row[17] else None
                prioridad = row[18] if row[18] else 'Media'  # Default a 'Media' si no se proporciona
                
                if not all([tipo_documento, num_documento, genero, nombres, apellido_paterno, 
                           fecha_nacimiento, estado_civil, email, celular, direccion, 
                           distrito_nombre]):
                    errores.append(f"Fila {row_idx}: Faltan campos requeridos")
                    continue
                
                # Obtener ID del departamento
                id_departamento = None
                lookup_conn = None
                try:
                    lookup_conn = get_db_connection()
                    if lookup_conn:
                        lookup_cur = lookup_conn.cursor(dictionary=True)
                        lookup_cur.execute("SELECT id_departamento FROM TblDepartamentos WHERE nombre = %s", (departamento_nombre,))
                        dept_result = lookup_cur.fetchone()
                        lookup_cur.close()
                        
                        if not dept_result:
                            errores.append(f"Fila {row_idx}: Departamento '{departamento_nombre}' no encontrado")
                            continue
                        id_departamento = dept_result['id_departamento']
                    else:
                        errores.append(f"Fila {row_idx}: Error de conexión buscando departamento")
                        continue
                except Exception as e:
                    logger.error(f"Error buscando departamento (fila {row_idx}): {e}")
                    errores.append(f"Fila {row_idx}: Error buscando departamento - {str(e)}")
                    continue
                finally:
                    if lookup_conn:
                        try:
                            lookup_conn.close()
                        except:
                            pass
                
                # Obtener ID de la provincia
                id_provincia = None
                lookup_conn = None
                try:
                    lookup_conn = get_db_connection()
                    if lookup_conn:
                        lookup_cur = lookup_conn.cursor(dictionary=True)
                        lookup_cur.execute(
                            "SELECT id_provincia FROM TblProvincias WHERE nombre = %s AND id_departamento = %s", 
                            (provincia_nombre, id_departamento)
                        )
                        prov_result = lookup_cur.fetchone()
                        lookup_cur.close()
                        
                        if not prov_result:
                            errores.append(f"Fila {row_idx}: Provincia '{provincia_nombre}' no encontrada en {departamento_nombre}")
                            continue
                        id_provincia = prov_result['id_provincia']
                    else:
                        errores.append(f"Fila {row_idx}: Error de conexión buscando provincia")
                        continue
                except Exception as e:
                    logger.error(f"Error buscando provincia (fila {row_idx}): {e}")
                    errores.append(f"Fila {row_idx}: Error buscando provincia - {str(e)}")
                    continue
                finally:
                    if lookup_conn:
                        try:
                            lookup_conn.close()
                        except:
                            pass
                
                # Obtener ID del distrito - ahora con el id_provincia
                id_distrito = None
                lookup_conn = None
                try:
                    lookup_conn = get_db_connection()
                    if lookup_conn:
                        lookup_cur = lookup_conn.cursor(dictionary=True)
                        lookup_cur.execute(
                            "SELECT id_distrito FROM TblDistritos WHERE nombre = %s AND id_provincia = %s", 
                            (distrito_nombre, id_provincia)
                        )
                        dist_result = lookup_cur.fetchone()
                        lookup_cur.close()
                        
                        if not dist_result:
                            errores.append(f"Fila {row_idx}: Distrito '{distrito_nombre}' no encontrado en {provincia_nombre}, {departamento_nombre}")
                            continue
                        id_distrito = dist_result['id_distrito']
                    else:
                        errores.append(f"Fila {row_idx}: Error de conexión buscando distrito")
                        continue
                except Exception as e:
                    logger.error(f"Error buscando distrito (fila {row_idx}): {e}")
                    errores.append(f"Fila {row_idx}: Error buscando distrito - {str(e)}")
                    continue
                finally:
                    if lookup_conn:
                        try:
                            lookup_conn.close()
                        except:
                            pass
                
                # Obtener IDs opcionales - cada uno en su CONEXIÓN SEPARADA
                # CAMBIO: Ahora pasamos los NOMBRES directamente al SP, no los IDs
                nombre_fuente_contacto = fuente_contacto  # Pasar el nombre tal cual
                nombre_estado_prospeccion = estado_prospeccion  # Pasar el nombre tal cual
                nombre_tipo_compra = tipo_compra  # Pasar el nombre tal cual
                
                # EJECUTAR SP - CONEXIÓN SEPARADA
                sp_conn = None
                sp_cur = None
                try:
                    sp_conn = get_db_connection()
                    if not sp_conn:
                        errores.append(f"Fila {row_idx}: Error de conexión para SP")
                        logger.error(f"No se pudo conectar a BD para SP en fila {row_idx}")
                        continue
                    
                    sp_cur = sp_conn.cursor()
                    
                    # Ejecutar SP
                    sp_cur.callproc('sp_InsertarClienteMasivo', [
                        num_documento,
                        tipo_documento,
                        nombres,
                        apellido_paterno,
                        apellido_materno,
                        fecha_nacimiento,
                        genero,
                        estado_civil,
                        email,
                        celular,
                        direccion,
                        id_distrito,
                        num_documento_asesor,
                        nombre_fuente_contacto,  # Pasar NOMBRE, no ID
                        proyecto_interes,  # Pasar NOMBRE, no ID
                        nombre_estado_prospeccion,  # Pasar NOMBRE, no ID
                        nombre_tipo_compra,  # Pasar NOMBRE, no ID
                        prioridad  # Agregar prioridad
                    ])
                    
                    # Consumir TODOS los resultados - ESTO ES CRÍTICO
                    mensaje_error = None
                    for result_set in sp_cur.stored_results():
                        row_result = result_set.fetchone()
                        if row_result:
                            exito = row_result[1] if len(row_result) > 1 else 0
                            if exito == 1:
                                clientes_importados += 1
                                logger.info(f"Fila {row_idx}: Cliente importado exitosamente")
                            else:
                                mensaje_error = row_result[0] if len(row_result) > 0 else 'Error desconocido'
                                errores.append(f"Fila {row_idx}: {mensaje_error}")
                                logger.warning(f"Fila {row_idx}: {mensaje_error}")
                    
                except Exception as sp_error:
                    errores.append(f"Fila {row_idx}: Error ejecutando SP - {str(sp_error)}")
                    logger.error(f"Error SP en fila {row_idx}: {sp_error}", exc_info=True)
                
                finally:
                    # Limpiar cursor y conexión del SP - DEBE ser en finally
                    if sp_cur:
                        try:
                            while sp_cur.nextset():
                                pass
                        except Exception as cleanup_error:
                            logger.debug(f"Cleanup error (ignorado): {cleanup_error}")
                        
                        try:
                            sp_cur.close()
                        except Exception as close_error:
                            logger.debug(f"Close error (ignorado): {close_error}")
                    
                    if sp_conn:
                        try:
                            sp_conn.close()
                        except Exception as conn_close_error:
                            logger.debug(f"Connection close error (ignorado): {conn_close_error}")
                
            except Exception as row_error:
                errores.append(f"Fila {row_idx}: Error procesando fila - {str(row_error)}")
                logger.error(f"Error en fila {row_idx}: {row_error}", exc_info=True)
        
        respuesta = {
            'success': True,
            'clientes_importados': clientes_importados,
            'total_filas_procesadas': ws.max_row - 1,
            'errores': errores
        }
        
        if errores:
            respuesta['message'] = f"Se importaron {clientes_importados} clientes con {len(errores)} error(es)"
            respuesta['success'] = clientes_importados > 0
        else:
            respuesta['message'] = f"Se importaron exitosamente {clientes_importados} clientes"
        
        logger.info(f"Importación completada: {clientes_importados} exitosos, {len(errores)} errores")
        return respuesta, 200 if respuesta['success'] else 400
        
    except Exception as e:
        from flask import current_app
        current_app.logger.error(f"Error general en importar_clientes_api: {e}", exc_info=True)
        return {'success': False, 'message': f'Error: {str(e)}'}, 500
